"""Simplifi P&L report parser."""
import openpyxl
import pandas as pd
from pathlib import Path
from typing import Dict, Any, Optional
import logging


logger = logging.getLogger(__name__)


class SimplifiParser:
    """Parser for Simplifi P&L export files (CSV/XLSX)."""

    def __init__(self):
        self.data: Optional[pd.DataFrame] = None
        self.monthly_summary: Optional[pd.DataFrame] = None

    def load_file(self, file_path: Path) -> pd.DataFrame:
        """Load Simplifi export file."""
        try:
            if file_path.suffix.lower() == '.csv':
                raise ValueError(f"Switching to only xlsx, not supported: {file_path.suffix}")
                # df = pd.read_csv(file_path)
            elif file_path.suffix.lower() in ['.xlsx', '.xls']:
                # Load the workbook and select the active sheet to get outline level
                workbook = openpyxl.load_workbook(file_path)
                sheet = workbook.active
                # Create an empty list to hold the levels
                outline_levels = []

                # Loop ONLY through the rows that contain data to build outline level column
                # Data starts on Excel row 2, after title, min_row=2 skips the title
                # sheet.iter_rows() is efficient for this.
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                    # The first cell in the row tuple gives us access to its properties
                    level = sheet.row_dimensions[row[0].row].outline_level
                    outline_levels.append(level)

                # Read the dataframe grid
                df = pd.read_excel(file_path)

                # Add the hierarchy level column
                df['HierarchyLevel'] = outline_levels

            else:
                raise ValueError(f"Unsupported file format: {file_path.suffix}")

            self.data = self._clean_data(df)
            return self.data

        except Exception as e:
            logger.error(f"Error loading Simplifi file: {e}")
            raise

    def _clean_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """Clean and standardize the imported data."""
        df_clean = df.copy()

        # 2) Normalize numeric-looking object columns:
        #    - remove leading "\'" or "'" artifacts
        #    - remove thousands separators
        #    - convert (123.45) -> -123.45
        #    - coerce to numeric
        obj_cols = df_clean.select_dtypes(include=['object']).columns.tolist()

        def to_numeric_series(s: pd.Series) -> pd.Series:
            # Work on string view
            s_str = s.astype(str).str.strip()

            # Strip leading \" and ' characters that may prefix negatives, e.g., "\'-315.87"
            s_str = s_str.str.replace(r"^[\\']+", "", regex=True)

            # Remove common currency/spacing artifacts (keep digits, minus, dot, parentheses, comma)
            # First handle parentheses negatives
            s_str = s_str.str.replace(r"^\((.*)\)$", r"-\1", regex=True)

            # Drop commas as thousand separators
            s_str = s_str.str.replace(",", "", regex=False)

            # Finally, to_numeric
            return pd.to_numeric(s_str, errors='coerce')

        for col in obj_cols:
            parsed = to_numeric_series(df_clean[col])
            # Only convert the column if a significant portion parsed as numbers
            # (tune threshold as needed)
            if parsed.notna().mean() >= 0.6:
                df_clean[col] = parsed

        # 2) Replace NaN with 0 only for numeric columns
        num_cols = df_clean.select_dtypes(include=['number']).columns
        df_clean[num_cols] = df_clean[num_cols].fillna(0)

        return df_clean

    def get_monthly_summary(self, start_date: str = None, end_date: str = None) -> pd.DataFrame:
        """Generate monthly summary of income/expenses by category."""
        if self.data is None:
            raise ValueError("No data loaded. Call load_file() first.")

        # TODO: Implement monthly aggregation logic
        # This is a placeholder implementation
        summary = self.data.copy()

        if start_date or end_date:
            # Filter by date range if provided
            pass

        self.monthly_summary = summary
        return summary

    def get_category_breakdown(self) -> Dict[str, Dict[str, float]]:
        """Get breakdown of expenses by category and subcategory."""
        if self.data is None:
            raise ValueError("No data loaded. Call load_file() first.")

        # TODO: Implement category breakdown logic
        return {}

    def export_processed_data(self, output_path: Path) -> None:
        """Export processed data to CSV."""
        if self.data is None:
            raise ValueError("No data loaded. Call load_file() first.")

        self.data.to_csv(output_path, index=False)
